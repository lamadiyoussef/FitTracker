from tkinter import *
from tkinter import ttk
import hashlib
from tkinter import messagebox
import os
import openpyxl




def enter_data():
    accepted = accept_var.get()

    if accepted == "Accepted":
        # required user info
        username = username_entry.get()
        email = email_entry.get()
        password = password_entry.get()
        confirm_password = confirm_password_entry.get()
        if password != confirm_password:
            messagebox.showwarning(title="Error", message="Passwords do not match. Please make sure your password and confirm password entries are identical.")
        elif email and password and username:
            # Check uniqueness of username and email
            filepath = "..\\DB.xlsx"
            if not os.path.exists(filepath):
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                heading = ["NOM", "PRENOM", "GENDER", "AGE", "HEIGHT", "WEIGHT", "USERNAME", "PASSWORD", "EMAIL",
                           "EMAIL_NOTIFICATIONS"]
                sheet.append(heading)
                workbook.save(filepath)


            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active
            usernames = [cell.value for cell in sheet['G'][1:]]  # Assuming username is in column E
            emails = [cell.value for cell in sheet['I'][1:]]  # Assuming email is in column G

            if username in usernames:
                messagebox.showwarning(title="Error", message="Username already exists. Please choose another one.")
                return
            if email in emails:
                messagebox.showwarning(title="Error", message="Email already exists. Please use a different one.")
                return

             # Hashing password
            hashed_password = hashlib.sha256(password.encode()).hexdigest()

            #other user info
            firstname = first_name_entry.get()
            lastname = last_name_entry.get()
            gender = sex_combobox.get()
            age = age_spinbox.get()

            height = height_entry.get()
            weight = weight_entry.get()


            # Check if height is a positive number
            try:
                height_value = float(height)
                if height_value <= 0:
                    messagebox.showwarning(title="Error", message="Height must be a positive number.")
                    return
            except ValueError:
                messagebox.showwarning(title="Error", message="Height must be a number.")
                return

            # Check if weight is a positive number
            try:
                weight_value = float(weight)
                if weight_value <= 0:
                    messagebox.showwarning(title="Error", message="Weight must be a positive number.")
                    return

            except ValueError:
                messagebox.showwarning(title="Error", message="Weight must be a number.")
                return

            filepath = "..\\DB.xlsx"

            if not os.path.exists(filepath):
                 workbook = openpyxl.Workbook()
                 sheet = workbook.active
                 heading = ["NOM", "PRENOM", "GENDER", "AGE", "HEIGHT", "WEIGHT", "USERNAME","PASSWORD", "EMAIL", "EMAIL_NOTIFICATIONS"]
                 sheet.append(heading)
                 workbook.save(filepath)
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active
            sheet.append([firstname, lastname, gender, age,height,weight, username, hashed_password, email, "on"])
            workbook.save(filepath)
            messagebox.showinfo(title="Success", message="You registered successfully!")

        else:
            messagebox.showwarning(title="Error", message="username, password and email are required.")
    else:
        messagebox.showwarning(title="Error", message="You have not accepted the terms")


window = Tk()
window.title("Sign Up")
window.geometry('340x440')

#window.geometry("500x400")

# c=Canvas(window,bg="gray16", height=200,width=200)
# filename=PhotoImage(file="img8.png")
# background_label=Label(window,image=filename)
# background_label.place(x=0,y=0,relwidth=1,relheight=1)
# c.pack()


frame = Frame(window,bg='#333333',width=600, height=500)
#frame.pack()
frame.place(relx=0.5, rely=0.5, anchor="center")

# Saving User Info
user_info_frame = LabelFrame(frame, text="User Information",bg='#333333')
user_info_frame.grid(row=0, column=0, padx=20, pady=10)

first_name_label = Label(user_info_frame, text="First Name")
first_name_label.grid(row=0, column=0)
last_name_label = Label(user_info_frame, text="Last Name")
last_name_label.grid(row=0, column=1)

first_name_entry = Entry(user_info_frame)
last_name_entry = Entry(user_info_frame)
first_name_entry.grid(row=1, column=0)
last_name_entry.grid(row=1, column=1)

age_label = Label(user_info_frame, text="Age")
age_spinbox = Spinbox(user_info_frame, from_=0, to=110)
age_label.grid(row=2, column=0)
age_spinbox.grid(row=3, column=0)

# Create labels and entries for weight, height
weight_label = Label(user_info_frame, text="Weight (kg):")
weight_entry = Entry(user_info_frame)
height_label = Label(user_info_frame, text="Height (cm):")
height_entry = Entry(user_info_frame)

# Arrange elements using grid layout
height_label.grid(row=2, column=1)
height_entry.grid(row=3, column=1)
weight_label.grid(row=2, column=2)
weight_entry.grid(row=3, column=2)

sex_label = Label(user_info_frame, text="Gender")
sex_combobox = ttk.Combobox(user_info_frame, values=["", "Male", "Female"])
sex_label.grid(row=0, column=2)
sex_combobox.grid(row=1, column=2)

username_label = Label(user_info_frame, text="Username")
username_label.grid(row=4,column=0)
username_entry = Entry(user_info_frame)
username_entry.grid(row=5 , column=0 )

password_label = Label(user_info_frame, text="Password")
password_label.grid(row=4,column=1)
password_entry = Entry(user_info_frame,show="*")
password_entry.grid(row=5 , column=1 )

confirm_password_label = Label(user_info_frame, text="Confirm Password")
confirm_password_label.grid(row=4,column=2)
confirm_password_entry = Entry(user_info_frame, show="*")
confirm_password_entry.grid(row=5 , column=2 )

email_label = Label(user_info_frame, text="Email Address")
email_label.grid(row=8,column=0)
email_entry = Entry(user_info_frame)
email_entry.grid(row=9 , column=0 )


for widget in user_info_frame.winfo_children():
    widget.grid_configure(padx=10, pady=5)


# Accept terms
terms_frame = LabelFrame(frame, text="Terms & Conditions")
terms_frame.grid(row=2, column=0, sticky="news", padx=20, pady=10)

accept_var = StringVar(value="Not Accepted")
terms_check = Checkbutton(terms_frame, text="I accept the terms and conditions.",variable=accept_var, onvalue="Accepted", offvalue="Not Accepted")
terms_check.grid(row=0, column=0)

# Button
button = Button(frame, text="Sign up", command=enter_data)
button.grid(row=3, column=0, sticky="news", padx=20, pady=10)

window.mainloop()